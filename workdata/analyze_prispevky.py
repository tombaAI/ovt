#!/usr/bin/env python3
"""
Hlubší analýza OVT Příspěvky.xlsx — inferuje roční nastavení příspěvků
a unikátní sadu členů přes všechny roky.
"""
import sys
from collections import Counter
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    import openpyxl

XLSX = Path(__file__).parent.parent / "zadani" / "OVT Příspěvky.xlsx"
OUT  = Path(__file__).parent / "prispevky_analysis.txt"

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
lines = []

YEAR_SHEETS = [s for s in wb.sheetnames if s.startswith("Příspěvky")]
lines.append(f"Listy: {YEAR_SHEETS}\n")

def to_int(v):
    if v is None: return None
    try: return int(float(v))
    except: return None

def mode_of(vals):
    vals = [v for v in vals if v is not None and v != 0]
    if not vals: return None
    return Counter(vals).most_common(1)[0][0]

all_members = {}   # member_number → latest data
year_stats  = {}   # year → {field: [values]}

for sheet_name in YEAR_SHEETS:
    year = int(sheet_name.split()[-1])
    ws   = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # najdi header řádek
    header_row = next((r for r in rows if r[0] == "ID"), None)
    if not header_row:
        lines.append(f"\n{sheet_name}: header nenalezen")
        continue
    header = list(header_row)
    hi = {v: i for i, v in enumerate(header) if v is not None}

    data = [r for r in rows[rows.index(header_row)+1:] if r[0] is not None]

    stats = {
        "count": len(data),
        "zakladni": [], "lod1": [], "lod2": [], "lod3": [],
        "brigada": [], "s_vybor": [], "s_tom": [], "s_ind": [],
    }

    def g(row, key):
        idx = hi.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    for row in data:
        mid = to_int(g(row, "ID"))
        if mid is None:
            continue

        member = {
            "member_number":    mid,
            "user_login":       g(row, "user_login"),
            "email":            g(row, "user_email"),
            "phone":            g(row, "tel"),
            "full_name":        g(row, "jmeno"),
            "variable_symbol":  to_int(g(row, "variabilni_symbol")),
            "csk_number":       to_int(g(row, "CSK_cislo")),
            "clen_ovt":         g(row, "clen_ovt"),
        }
        # zachovej nejnovější verzi
        if mid not in all_members or year >= all_members[mid]["_year"]:
            member["_year"] = year
            all_members[mid] = member

        base  = to_int(g(row, "zakladni_prispevky"))
        lod1  = to_int(g(row, "lod_1"))
        lod2  = to_int(g(row, "lod_2"))
        lod3  = to_int(g(row, "lod_3"))
        brig  = to_int(g(row, "n_brigada"))
        vybor = to_int(g(row, "s_vybor"))
        tom   = to_int(g(row, "s_tom"))
        ind   = to_int(g(row, "s_individualni"))

        if base:  stats["zakladni"].append(base)
        if lod1:  stats["lod1"].append(lod1)
        if lod2:  stats["lod2"].append(lod2)
        if lod3:  stats["lod3"].append(lod3)
        if brig and brig > 0:  stats["brigada"].append(brig)
        if vybor: stats["s_vybor"].append(abs(vybor))
        if tom:   stats["s_tom"].append(abs(tom))
        if ind:   stats["s_ind"].append(abs(ind))

    year_stats[year] = stats

    lines.append(f"\n{'='*60}")
    lines.append(f"ROK {year}  |  {stats['count']} záznamů")
    lines.append(f"  zakladni_prispevky : mode={mode_of(stats['zakladni'])}  values={sorted(set(stats['zakladni']))}")
    lines.append(f"  lod_1              : mode={mode_of(stats['lod1'])}   values={sorted(set(stats['lod1']))}")
    lines.append(f"  lod_2              : mode={mode_of(stats['lod2'])}   values={sorted(set(stats['lod2']))}")
    lines.append(f"  lod_3              : mode={mode_of(stats['lod3'])}   values={sorted(set(stats['lod3']))}")
    lines.append(f"  n_brigada (>0)     : mode={mode_of(stats['brigada'])}  values={sorted(set(stats['brigada']))}")
    lines.append(f"  s_vybor (abs)      : mode={mode_of(stats['s_vybor'])}  values={sorted(set(stats['s_vybor']))}")
    lines.append(f"  s_tom (abs)        : mode={mode_of(stats['s_tom'])}  values={sorted(set(stats['s_tom']))}")
    lines.append(f"  s_ind (abs)        : mode={mode_of(stats['s_ind'])}  values={sorted(set(stats['s_ind']))}")

lines.append(f"\n{'='*60}")
lines.append(f"CELKEM UNIKÁTNÍCH ČLENŮ: {len(all_members)}")
active = [m for m in all_members.values() if str(m.get('clen_ovt','')).lower() in ('ano',)]
lines.append(f"Aktivních (clen_ovt=Ano v posledním roce): {len(active)}")

OUT.write_text("\n".join(lines), encoding="utf-8")
print(f"Hotovo → {OUT}")
