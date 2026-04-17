#!/usr/bin/env python3
"""
Parsuje Excel soubory ze zadani/ a vypisuje strukturu + ukázková data.
Výstup ukládá do workdata/xlsx_analysis.txt
"""
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instaluji openpyxl...")
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    import openpyxl

OUTPUT = Path(__file__).parent / "xlsx_analysis.txt"
ZADANI = Path(__file__).parent.parent / "zadani"

files = sorted(ZADANI.glob("*.xlsx"))

lines = []

for path in files:
    lines.append("=" * 70)
    lines.append(f"SOUBOR: {path.name}")
    lines.append("=" * 70)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"\n  --- List: {sheet_name} ---")

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            lines.append("  (prázdný list)")
            continue

        # záhlaví = první neprázdný řádek
        header_idx = next(
            (i for i, r in enumerate(rows) if any(c is not None for c in r)),
            None
        )
        if header_idx is None:
            lines.append("  (žádná data)")
            continue

        header = rows[header_idx]
        non_empty_cols = [i for i, v in enumerate(header) if v is not None]

        lines.append(f"  Sloupce ({len(non_empty_cols)}):")
        for i in non_empty_cols:
            lines.append(f"    [{i}] {header[i]}")

        # ukázka prvních 5 datových řádků
        data_rows = [r for r in rows[header_idx + 1:] if any(c is not None for c in r)]
        lines.append(f"\n  Počet datových řádků: {len(data_rows)}")
        lines.append("  Ukázka (prvních 5 řádků):")
        for row in data_rows[:5]:
            sample = {header[i]: row[i] for i in non_empty_cols if i < len(row)}
            lines.append(f"    {sample}")

    wb.close()

OUTPUT.write_text("\n".join(lines), encoding="utf-8")
print(f"Hotovo → {OUTPUT}")
