#!/usr/bin/env python3
"""
Generates import_data.sql from OVT Příspěvky.xlsx.
Covers all years found in the workbook (2020-2026).
Output: workdata/import_data.sql — apply via Neon SQL Editor.
"""
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    import openpyxl

XLSX = Path(__file__).parent.parent / "zadani" / "OVT Příspěvky.xlsx"
OUT  = Path(__file__).parent / "import_data.sql"

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
YEAR_SHEETS = [(int(s.split()[-1]), s) for s in wb.sheetnames if s.startswith("Příspěvky")]
YEAR_SHEETS.sort()

# ── Inferred period settings (from analyze_prispevky.py) ─────────────────────
PERIOD_SETTINGS = {
    2020: dict(amount_base=800,  amount_boat1=450, amount_boat2=200, amount_boat3=300, discount_committee=400, discount_tom=400, brigade_surcharge=400, due_date="2020-02-29"),
    2021: dict(amount_base=800,  amount_boat1=450, amount_boat2=200, amount_boat3=300, discount_committee=400, discount_tom=400, brigade_surcharge=400, due_date="2021-02-28"),
    2022: dict(amount_base=800,  amount_boat1=450, amount_boat2=200, amount_boat3=300, discount_committee=400, discount_tom=400, brigade_surcharge=400, due_date="2022-02-28"),
    2023: dict(amount_base=1000, amount_boat1=800, amount_boat2=500, amount_boat3=500, discount_committee=500, discount_tom=500, brigade_surcharge=500, due_date="2023-02-28"),
    2024: dict(amount_base=1000, amount_boat1=800, amount_boat2=500, amount_boat3=500, discount_committee=500, discount_tom=500, brigade_surcharge=500, due_date="2024-02-29"),
    2025: dict(amount_base=1000, amount_boat1=800, amount_boat2=500, amount_boat3=500, discount_committee=500, discount_tom=500, brigade_surcharge=500, due_date="2025-02-28"),
    2026: dict(amount_base=1500, amount_boat1=1200, amount_boat2=800, amount_boat3=800, discount_committee=500, discount_tom=500, brigade_surcharge=0,   due_date="2026-02-28"),
}

def q(v):
    """SQL-escape a value."""
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return str(int(v))
    s = str(v).replace("'", "''")
    return f"'{s}'"

def to_int(v):
    if v is None: return None
    try: return int(float(v))
    except: return None

def to_bool(v):
    if v is None: return None
    return str(v).strip().lower() in ("ano", "yes", "true", "1")

def to_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.strftime("%Y-%m-%d")
    try:
        return datetime.strptime(str(v)[:10], "%Y-%m-%d").strftime("%Y-%m-%d")
    except: return None

# ── Collect all members across years (latest year wins) ──────────────────────
all_members: dict[int, dict] = {}

for year, sheet_name in YEAR_SHEETS:
    ws   = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header_row = next((r for r in rows if r[0] == "ID"), None)
    if not header_row: continue
    hi = {v: i for i, v in enumerate(header_row) if v is not None}

    def g(row, key):
        idx = hi.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    data = [r for r in rows[rows.index(header_row)+1:] if r[0] is not None]
    for row in data:
        mid = to_int(g(row, "ID"))
        if mid is None: continue
        if mid not in all_members or year >= all_members[mid]["_year"]:
            all_members[mid] = {
                "_year":           year,
                "id":              mid,
                "user_login":      g(row, "user_login"),
                "email":           g(row, "user_email"),
                "phone":           g(row, "tel"),
                "full_name":       g(row, "jmeno") or f"Člen #{mid}",
                "variable_symbol": to_int(g(row, "variabilni_symbol")),
                "csk_number":      to_int(g(row, "CSK_cislo")),
                "is_active":       to_bool(g(row, "clen_ovt")) or False,
            }

# ── Collect contributions ─────────────────────────────────────────────────────
contributions: list[dict] = []   # list of {year, member_id, ...}

for year, sheet_name in YEAR_SHEETS:
    ws   = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header_row = next((r for r in rows if r[0] == "ID"), None)
    if not header_row: continue
    hi = {v: i for i, v in enumerate(header_row) if v is not None}

    def g(row, key):
        idx = hi.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    data = [r for r in rows[rows.index(header_row)+1:] if r[0] is not None]
    for row in data:
        mid = to_int(g(row, "ID"))
        if mid is None: continue

        # total: use dedicated total column if present, else sum parts
        total_col = None
        for col in (f"{year-1}_Prispevky", f"{year}_Prispevky", "celkem", "celkem_prispevky"):
            if col in hi:
                total_col = to_int(g(row, col))
                break

        contributions.append({
            "year":                 year,
            "member_id":            mid,
            "amount_total":         total_col,
            "amount_base":          to_int(g(row, "zakladni_prispevky")),
            "amount_boat1":         to_int(g(row, "lod_1")),
            "amount_boat2":         to_int(g(row, "lod_2")),
            "amount_boat3":         to_int(g(row, "lod_3")),
            "discount_committee":   to_int(g(row, "s_vybor")),
            "discount_tom":         to_int(g(row, "s_tom")),
            "discount_individual":  to_int(g(row, "s_individualni")),
            "brigade_surcharge":    to_int(g(row, "n_brigada")),
            "paid_amount":          to_int(g(row, "kolik")),
            "paid_at":              to_date(g(row, "kdy")),
            "is_paid":              to_bool(g(row, "Zaplaceno")),
            "note":                 g(row, "Poznámka") or g(row, "poznamka_k_platbe"),
        })

# ── Generate SQL ──────────────────────────────────────────────────────────────
lines = []
lines.append("-- OVT import: members, contribution_periods, member_contributions")
lines.append(f"-- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
lines.append("-- Apply via Neon SQL Editor\n")

lines.append("BEGIN;\n")

# 1. members
lines.append("-- ── Members ─────────────────────────────────────────────────")
for m in sorted(all_members.values(), key=lambda x: x["id"]):
    lines.append(
        f"INSERT INTO app.members (id, user_login, email, phone, full_name, variable_symbol, csk_number, is_active) VALUES ("
        f"{q(m['id'])}, {q(m['user_login'])}, {q(m['email'])}, {q(m['phone'])}, "
        f"{q(m['full_name'])}, {q(m['variable_symbol'])}, {q(m['csk_number'])}, {q(m['is_active'])}"
        f") ON CONFLICT (id) DO UPDATE SET "
        f"user_login={q(m['user_login'])}, email={q(m['email'])}, phone={q(m['phone'])}, "
        f"full_name={q(m['full_name'])}, variable_symbol={q(m['variable_symbol'])}, "
        f"csk_number={q(m['csk_number'])}, is_active={q(m['is_active'])}, updated_at=now();"
    )

lines.append("")

# 2. contribution_periods
lines.append("-- ── Contribution periods ────────────────────────────────────")
for year, s in sorted(PERIOD_SETTINGS.items()):
    if year not in [y for y, _ in YEAR_SHEETS]:
        continue
    dd = q(s["due_date"])
    lines.append(
        f"INSERT INTO app.contribution_periods "
        f"(year, amount_base, amount_boat1, amount_boat2, amount_boat3, "
        f"discount_committee, discount_tom, brigade_surcharge, due_date) VALUES ("
        f"{year}, {s['amount_base']}, {s['amount_boat1']}, {s['amount_boat2']}, {s['amount_boat3']}, "
        f"{s['discount_committee']}, {s['discount_tom']}, {s['brigade_surcharge']}, {dd}"
        f") ON CONFLICT (year) DO UPDATE SET "
        f"amount_base={s['amount_base']}, amount_boat1={s['amount_boat1']}, "
        f"amount_boat2={s['amount_boat2']}, amount_boat3={s['amount_boat3']}, "
        f"discount_committee={s['discount_committee']}, discount_tom={s['discount_tom']}, "
        f"brigade_surcharge={s['brigade_surcharge']}, due_date={dd};"
    )

lines.append("")

# 3. member_contributions
lines.append("-- ── Member contributions ────────────────────────────────────")
for c in sorted(contributions, key=lambda x: (x["year"], x["member_id"])):
    year = c["year"]
    if year not in PERIOD_SETTINGS:
        continue
    # note can be long — truncate at 500 chars
    note = c.get("note")
    if note and len(str(note)) > 500:
        note = str(note)[:500]

    lines.append(
        f"INSERT INTO app.member_contributions "
        f"(member_id, period_id, amount_total, amount_base, amount_boat1, amount_boat2, amount_boat3, "
        f"discount_committee, discount_tom, discount_individual, brigade_surcharge, "
        f"paid_amount, paid_at, is_paid, note) "
        f"SELECT {q(c['member_id'])}, id, "
        f"{q(c['amount_total'])}, {q(c['amount_base'])}, {q(c['amount_boat1'])}, "
        f"{q(c['amount_boat2'])}, {q(c['amount_boat3'])}, "
        f"{q(c['discount_committee'])}, {q(c['discount_tom'])}, {q(c['discount_individual'])}, "
        f"{q(c['brigade_surcharge'])}, {q(c['paid_amount'])}, {q(c['paid_at'])}, "
        f"{q(c['is_paid'])}, {q(note)} "
        f"FROM app.contribution_periods WHERE year={year} "
        f"ON CONFLICT (member_id, period_id) DO UPDATE SET "
        f"amount_total={q(c['amount_total'])}, paid_amount={q(c['paid_amount'])}, "
        f"paid_at={q(c['paid_at'])}, is_paid={q(c['is_paid'])};"
    )

lines.append("")
lines.append("COMMIT;")
lines.append("")

# summary
total_members = len(all_members)
active_members = sum(1 for m in all_members.values() if m["is_active"])
lines.append(f"-- Summary: {total_members} members ({active_members} active), "
             f"{len(contributions)} contribution records across {len(YEAR_SHEETS)} years")

OUT.write_text("\n".join(lines), encoding="utf-8")
print(f"Hotovo → {OUT}")
print(f"  {total_members} členů ({active_members} aktivních)")
print(f"  {len(contributions)} záznamy příspěvků ({len(YEAR_SHEETS)} let)")
